import re
import datetime
from django.utils import timezone
from django.shortcuts import render, redirect
from django.utils.dateparse import parse_date
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail, EmailMultiAlternatives
from userapp.util import *
# from userapp.util import mail_to_bm, app_task, leave_app_mail, leave_mail_approved
from userapp.models import EmployeeMaster, DesignationMaster, BranchMaster, DivisionMaster
from .models import LeaveMaster,CoverageMaster,LeaveTypesMaster,StateName,HolidayListMaster,HolidayTypeMaster,BalancedLeaveMaster, LeaveAttendanceLog
from datetime import date, timedelta
from django.core.files.storage import FileSystemStorage
# from django.core.files.storage import FileSystemStorage
from django.http import JsonResponse,HttpResponse
import csv
import openpyxl                            # Openpyxl package should be installed in system
from openpyxl import load_workbook
from string import Template
from django.http import HttpResponse
from django.contrib.auth.models import User

from django.shortcuts import render_to_response
from django.http import HttpResponseBadRequest
from django.template import RequestContext
import xlrd                                 # xlrd package should be installed in system

# from django.core.exceptions import ValidationError
# from django.utils.translation import gettext_lazy as _

# Create your views here.

def leave_index(request):
  """
  to show all employee records
  """
  try:
    dates = []
    dic = {}
    leave_branch_name = ''
    value = ''
    date_from = ''
    date_till = ''
    dates_list_to_pass = []
    atdn_list =[] 
    date_list =[]
    date_1 = datetime.date.today
    lv_date_from = datetime.date.today
    lv_date_to = datetime.date.today
    emps = EmployeeMaster.objects.all()
    leavemaster = LeaveMaster.objects.all()
    lev_typ = LeaveTypesMaster.objects.all()
    branch_list = BranchMaster.objects.all()
    holiday_list = HolidayListMaster.objects.all()
    
    if request.method == 'POST':
      """
      Search function for employee list view employee code,branch and status
      """
      emp_br = request.POST.getlist('leave_branch',False)
      date_from = request.POST.get('date_from',False)
      date_till = request.POST.get('date_to',False)
      name =''
      d1 = datetime.datetime.strptime(date_from, '%Y-%m-%d').date()            # From date of Attendance Report
      d2 = datetime.datetime.strptime(date_till, '%Y-%m-%d').date()            # Till date of Attendance Report
      delta = d2 - d1                                              # timedelta
      if emp_br :
        if emp_br[0]!="":                                                      #Check if branch is selected in Attendance report
          leave_branch = branch_list.get(id=emp_br[0])
          leave_branch_name = leave_branch.br_name
          emps_br_filter = emps.filter(master_table_type='BRANCH')             #Get all Employees Of Selected Branch
          if emps_br_filter: 
            emps_br_filter_2 = emps.filter(master_table_id=emp_br[0])
            if emps_br_filter_2:

              #Get all Employees Of Selected Branch

              if emps_br_filter_2 != None :
                for emp in emps_br_filter_2:
                  name = emp.emp_fname + emp.emp_lname
                  leaves = leavemaster.filter(emp_id = emp)                     #Check if Employee has taken any leave in between selected date range

                  if len(leaves) > 0:
                    date_list = []
                    dates = []
                    dates_list_to_pass = []
                    for i in range(delta.days + 1):                             #Loop to get Date 1 by 1 which is btw From Date and Till Date Selected by User
                      date_1 = d1 + timedelta(days=i)                           # date_1 is first date of difference(delta = d1 - d2)
                      # dates.append(date_1)
                      date_val = str(date_1)                                    # Converting Date to String
                      date_val1 = date_val.split('-')
                      # if dic != {}:
                      #   dates_list_to_pass = []
                      if dates_list_to_pass == [] :                             # Check if List is empty,to avoid date repeatation
                        dates.append(date_val1[2])                              # Appending all dates between From and To Date Selected 
                        day = date_1.weekday()                                  # Checking if particular date is Sunday
                        for leave in leaves:
                          lv_date_from = leave.date_from.date()                     #Employee from date leave
                          lv_date_to = leave.date_to.date()                         #Employee to date leave
                          delta_2 = lv_date_to - lv_date_from  
                          
                          for hol_list in holiday_list:
                            hol_list_date = str(hol_list.holiday_date).split(' ')
                            hol_list_date_str = datetime.datetime.strptime(str(hol_list_date[0]), '%Y-%m-%d').date()
                            # hol_list_date_str = datetime.datetime.strptime(str('2018-05-01'), '%Y-%m-%d').date()   #
                            
                            for lev_date in range(delta_2.days + 1):                  # loop to check if current leave date is equal to current(Date in between the provided date range) date
                              date_lev = lv_date_from + timedelta(days=lev_date)
                              if date_1 ==hol_list_date_str:
                              # if str(date_1) == '2018-05-01' :
                                value = hol_list.holiday_type_id.short_name_holiday_type
                                break
                                # value = hol_list.holiday_type_id.short_name_holiday_type
                              elif day == 6:
                                value = 'WO'  
                                break                   
                              elif str(date_lev) == str(date_1):                                  # Comparing if leave from date is equal to current (date selected in attendance report)date
                                value = leave.leave_type_id.short_name_leave_type
                                break
                              else:
                                pass
                            if value == '':
                              value = 'P'
                        date_list.append(value)
                        value = ''
                    dates_list_to_pass = dates
                    dic = {'emp_code':emp.emp_code,'emp_name':name,'dates':date_list}
                    atdn_list.append(dic)


    else:
      leaves = LeaveMaster.objects.all()
    desgn_list = DesignationMaster.objects.all()
    div_list = DivisionMaster.objects.all()
    
    
    return render(request, 'attendanceapp/leave_list.html', {'leave_branch':leave_branch_name,'date_from':date_from,'date_to':date_till,'branch_list':branch_list,'dates':dates_list_to_pass,'atdn_list':atdn_list})

  except Exception as e: 
    return render(request, 'userapp/page_404.html')

def export_to_excel(request):
  """
  Export To CSV File Function
  """
  dates = []
  header_exl_data = []
  dates_list_to_pass = []
  employees_attendance_list_exl_data = []
  dic = {}
  atdn_list =[] 
  date_list =[]
  date_1 = datetime.date.today
  lv_date_from = datetime.date.today
  lv_date_to = datetime.date.today
  # try:
  emps = EmployeeMaster.objects.all()
  leavemaster = LeaveMaster.objects.all()
  lev_typ = LeaveTypesMaster.objects.all()
  branch_list = BranchMaster.objects.all()
  holiday_list = HolidayListMaster.objects.all()

  
  if request.method == 'POST':

    """
    Search function for employee list view employee code,branch and status
    """
    emp_br = request.POST.get('leave_branch_1',False)
    date_from = request.POST.get('date_from',False)
    date_till = request.POST.get('date_to',False)
    name =''
    d1 = datetime.datetime.strptime(date_from, '%Y-%m-%d').date()            # From date of Attendance Report
    d2 = datetime.datetime.strptime(date_till, '%Y-%m-%d').date()            # Till date of Attendance Report
    delta = d2 - d1                                              # timedelta
    # if emp_br :
    if emp_br  !="":                                                      #Check if branch is selected in Attendance report
      leave_branch = branch_list.get(br_name=emp_br)
      
      emps_br_filter = emps.filter(master_table_type='BRANCH')             #Get all Employees Of Selected Branch
      if emps_br_filter: 
        emps_br_filter_2 = emps.filter(master_table_id=leave_branch.id)
        if emps_br_filter_2:

          #Get all Employees Of Selected Branch

          if emps_br_filter_2 != None :
            employee_code = 'Employee Code'
            employee_name = 'Employee Name'
            header_exl_data.append(employee_code)
            header_exl_data.append(employee_name)
            for emp in emps_br_filter_2:
              name = emp.emp_fname + emp.emp_lname
              leaves = leavemaster.filter(emp_id = emp)                     #Check if Employee has taken any leave in between selected date range
              
              if len(leaves) > 0:
                date_list = []
                dates = []
                dates_list_to_pass = []
                for i in range(delta.days + 1):                             #Loop to get Date 1 by 1 which is btw From Date and Till Date Selected by User
                  date_1 = d1 + timedelta(days=i)
                  # dates.append(date_1)
                  test = str(date_1)
                  test_1 = test.split('-')
                  if dates_list_to_pass == [] :
                    dates.append(test_1[2])
                    # header_exl_data.append(test_1[2])
                    day = date_1.weekday() 
                    for leave in leaves:
                      lv_date_from = leave.date_from.date()                     #Employee from date leave
                      lv_date_to = leave.date_to.date()                         #Employee to date leave
                      delta_2 = lv_date_to - lv_date_from  
                      for hol_list in holiday_list:
                        test_3 = str(hol_list.holiday_date).split(' ')
                        test_4 = datetime.datetime.strptime(str(hol_list_date[0]), '%Y-%m-%d').date()   #
                        
                        for lev_date in range(delta_2.days + 1):                  # loop to check if current leave date is equal to current(Date in between the provided date range) date
                          date_lev = lv_date_from + timedelta(days=lev_date)
                          if str(date_1) == hol_list_date_str:
                          # if str(date_1) == '2018-05-01' :
                              value = hol_list.holiday_type_id.short_name_holiday_type
                              break
                              # value = hol_list.holiday_type_id.short_name_holiday_type
                          elif day == 6:
                            value = 'WO'  
                            break                   
                          elif str(date_lev) == str(date_1):                                  # Comparing if leave from date is equal to current (date selected in attendance report)date
                            value = leave.leave_type_id.short_name_leave_type
                            break
                        if value == '':
                          value = 'P'
                  
                    date_list.append(value)
                    value = ''
                  header_exl_data_date= dates
                header_exl = header_exl_data + header_exl_data_date
                dic = {'emp_code':emp.emp_code,'emp_name':name,'dates':date_list}
                atdn_list.append(dic)
          else:
            msg='Something Went Wrong....There is no Employee Data Present of Selected Branch !!!'
            return render(request, 'attendanceapp/error.html',{'msg':msg})

  response = HttpResponse(content_type='application/vnd.ms-excel')
  response['Content-Disposition'] = 'attachment; filename="users.csv"'
  # print('---header_exl_data---',header_exl_data)
  writer = csv.writer(response)        # Create Blank CSV
  writer.writerow(['ATTENDANCE REPORT OF EMPLOYEES:'])
  writer.writerow(header_exl)     # ADD Header
  
  for a_list in atdn_list:             # Loop to add User Attendance Report Row
    emp_code = a_list['emp_code']
    emp_name = a_list['emp_name']
    employees_attendance_list_exl_data.append(emp_code)
    employees_attendance_list_exl_data.append(emp_name)
    for date in a_list['dates']:
      employees_attendance_list_exl_data.append(date)
    writer.writerow(employees_attendance_list_exl_data)
    employees_attendance_list_exl_data = []
  return response
  # except Exception as e: 
  #   print("=*****************Exception :",e)
  #   return redirect('leave_index')

 
def error_go_back(request):
  # Function to Redirect User to Leave List Home Page
  leave_list = LeaveMaster.objects.all()
  return render(request, 'attendanceapp/leave_application_list.html',{'leave_list':leave_list})


def import_holiday_list_excel(request):
  dic = {}
  if request.method == "POST":
    
    if request.FILES:
      upload = request.FILES['myfile']
      # content = upload.read()
      
      workbook = openpyxl.load_workbook(upload, read_only=True)   # To get Uploaded file name
      # Get name of the first sheet and then open sheet by name
      first_sheet = workbook.get_sheet_names()[0]
      worksheet = workbook.get_sheet_by_name(first_sheet)

      holiday_list_master = HolidayListMaster.objects.all()  
      holiday_list_master.delete()
      for row in worksheet.iter_rows(row_offset=1): # Offset for header # To add holiday list value one by one in database
        holiday_list = HolidayListMaster()
        # state_master = StateName.objects.all()
        # holiday_type_master = HolidayTypeMaster.objects.all()
        state_ins = StateName.objects.filter(id = row[5].value).first()
        holiday_type_ins = HolidayTypeMaster.objects.filter(id = row[6].value).first()

        
        holiday_list.holiday_name = str(row[1].value)  # Fetch values from excel sheet and add them in model field
        holiday_list.holiday_date = row[2].value
        holiday_list.holiday_day = row[3].value
        holiday_list.assigned_by = row[4].value
        holiday_list.state_id = state_ins
        holiday_list.holiday_type_id = holiday_type_ins
        holiday_list.coverage = row[7].value
        holiday_list.region = row[8].value
        holiday_list.save()                        # Save Holiday
      return render(request, 'attendanceapp/holiday_list.html')
    else:
      msg='Please Select Holiday List Excel Sheet to Upload !!!'
      # validate_even(value)
      return render(request, 'attendanceapp/error.html',{'msg':msg})
  else:
      msg='Something Went Wrong...Try again !!!'
      return render(request, 'attendanceapp/error.html',{'msg':msg})



# def validate_even(value):           # To Raise Validations (value-contains message to raise in validation popup)
#   if value:
#       raise ValueError(
#           _('%(value)s'),
#           params={'value': value},
#       )

def import_emp_balanced_leave_list_excel(request):        # To import balanced leave list of all employees
  dic = {}
  lev_type_list = []
  if request.method == "POST" :
    if request.FILES:
      upload = request.FILES['myfile']
      workbook = openpyxl.load_workbook(upload, read_only=True) # to get uploaded file value

      # Get name of the first sheet and then open sheet by name

      first_sheet = workbook.get_sheet_names()[0]
      worksheet = workbook.get_sheet_by_name(first_sheet)
      max_col = int(worksheet.max_column)-1
      i = 2                           # Supporting Variable to handle list which includes all leave type eg:['PL','CL','SL','MatL','PatL','OT','LWP'] 
      
      for row in worksheet.iter_rows(row_offset=0):

        #Get all employee with employee code in list and delete their Balanced Leaves Value and Add New Balanced Leave Of those Employees 
      
        get_emp_balanced_lev = BalancedLeaveMaster.objects.filter(emp_code=row[0].value)
        for emp_bal_lev in get_emp_balanced_lev:
          emp_bal_lev.delete()
        while i <= max_col:
          lev_type = row[i].value
          lev_type_list.append(lev_type)
          i =  i + 1

      for row in worksheet.iter_rows(row_offset=1): # Offset for header
        y = 2                                  # Supporting Variable to get list of balanced leave of respective leave type for each employee 
        for lev_type_list_val in lev_type_list:
          leave_type = LeaveTypesMaster.objects.get(short_name_leave_type=lev_type_list_val)
          y <= max_col
          balanced_list = BalancedLeaveMaster()
          balanced_list.emp_code = str(row[0].value)       # To get value from excel sheet one by one
          balanced_list.emp_name = str(row[1].value)
          balanced_list.balanced_leave_type_id = leave_type
          balanced_list.balanced_leave = row[y].value
          balanced_list.save()
          y = y + 1

      return render(request, 'attendanceapp/view_balanced_leave.html')
    else:
      msg='Please Select Balanced Leave Excel Sheet to Upload !!!'
      return render(request, 'attendanceapp/error.html',{'msg':msg})


def leave_application(request):
  """
  to send employee details to app admin to create app account
  """
  try:
    emps = EmployeeMaster.objects.all()
    leaves = LeaveTypesMaster.objects.all()

    return render(request, 'attendanceapp/leave_application_form.html',{'emps':emps,'leaves':leaves})
    # return render(request, 'userapp/form.html')
  except Exception as e: 
    return redirect('index')

def leave_application_list(request):
  """
  To List All Leave Applications
  """
  try:
    leave_list = LeaveMaster.objects.all()
    # leaves = LeaveTypesMaster.objects.all()

    return render(request, 'attendanceapp/leave_application_list.html',{'leave_list':leave_list})
    # return render(request, 'userapp/form.html')
  except Exception as e: 
    return redirect('index')
    

def add_holiday(request):
  """
  To Add Holiday List According to State
  """
  try:
    states = StateName.objects.all()
    holidays = HolidayListMaster.objects.all()
    holiday_type = HolidayTypeMaster.objects.all()
    coverages = CoverageMaster.objects.all()
    # leaves = LeaveMaster.objects.all()
    

    return render(request, 'attendanceapp/add_holiday.html',{'coverages':coverages,'states':states,'holidays':holidays,'holiday_type':holiday_type})
  except Exception as e: 
    return redirect('leave_index')

def add_holiday_type(request):
  """
  To Add Holiday Type 
  """
  try:
    # states = StateName.objects.all()
    holiday_type = HolidayTypeMaster.objects.all()

    return render(request, 'attendanceapp/add_holiday_type.html',{'holiday_type':holiday_type})
  except Exception as e: 
    return redirect('leave_index')


def submit_holiday_type(request):
  """
  To Submit Holiday Type in Database
  """
  try:
    if request.method=='POST':
      holiday_type = HolidayTypeMaster()
      hol_type = request.POST.get('hol_type',None)
      if hol_type is not None:
        holiday_type.holiday_type = hol_type
      
      holiday_type.save()


      msg = 'Holiday Type added Successfully !!!'
      return render(request, 'attendanceapp/thank_you.html',{'msg':msg})
  except Exception as e: 
    return redirect('leave_index')

def submit_new_holiday(request):
  """
  To Add New Holiday for Selected State in Database
  """
  try:
    if request.method=='POST':
      states = StateName.objects.all()
      holidays = HolidayListMaster()

      hol_name = request.POST.get('hol_name',False)
      hol_date = request.POST.get('hol_date',None)
      hol_day = request.POST.get('hol_day',False)
      region = request.POST.get('region',False)
      coverage = request.POST.get('coverage',False)
      assigned_by = request.POST.get('assigned_by',False)
      # fix_hol_date = request.POST.get('hol_date')
      # opt_hol_date = request.POST.get('opt_hol_date')
      state = request.POST.getlist('state',False)
      state_ins = StateName.objects.get(id = state[0])

      holiday_type = request.POST.getlist('holiday_type',False)
      holiday_type_ins = HolidayTypeMaster.objects.get(id = holiday_type[0])

      holidays.holiday_name = hol_name
      holidays.holiday_date = hol_date
      holidays.holiday_day = hol_day
      holidays.leave_type_id = holiday_type_ins
      holidays.coverage = coverage
      holidays.region = region
      holidays.assigned_by = assigned_by
      holidays.state_id=state_ins
      
      # if fix_hol_date is not None:
      #   holidays.fixed_holiday_date = fix_hol_date
      # if opt_hol_date is not None:
      #   holidays.optional_holiday_date=opt_hol_date
       
      
      holidays.save()

      msg = 'Holiday Added Successfully in Database!!!'
      return render(request, 'attendanceapp/thank_you.html',{'msg':msg})
  except Exception as e: 
    return redirect('leave_index')

def holiday_list_search_view(request):
  """
  To Direct to List/View Holidays where User can get Holiday list by Selecting State and hitting Search Button
  """
  try:
    states = StateName.objects.all()
    holidays = HolidayListMaster.objects.all()

    return render(request, 'attendanceapp/holiday_list.html',{'states':states,'holidays':holidays})
  except Exception as e: 
    return redirect('leave_index')

def list_holiday_index(request):
  """
  To List/View Holidays Data for Selected State
  """
  try:
    states = StateName.objects.all()
    holidays = HolidayListMaster.objects.all()

    return render(request, 'attendanceapp/holiday_list.html',{'states':states,'holidays':holidays})
  except Exception as e: 
    return redirect('leave_index')
  

def submit_leave(request):
  """
  OnClick of Submit button in Leave Application Form, mail will go to respective Reporting manager to approve or disapprove leave request
  """
  # try:
  if request.method == 'POST' :
    user_name = request.POST.get('user_name')
    emp_code = request.POST.get('emp_code')
    user_email = request.POST.get('user_email')
    leave_type = request.POST.get('leave_type')
    date_from = request.POST['date_from']
    date_to = request.POST['date_to']

    sick_lev_doc_pic = ''
    if 'myfile' in request.FILES:
      sick_lev_doc_pic = request.FILES['myfile']
      
      fs = FileSystemStorage()
      filename = fs.save(sick_lev_doc_pic.name, sick_lev_doc_pic)
      uploaded_file_url = fs.url(filename)

    leave_ins = LeaveMaster()
    emps = EmployeeMaster.objects.filter(emp_code = emp_code)

    if emps:
      for emp in emps:
        leave_type_ins = LeaveTypesMaster.objects.get(leave_type = leave_type)

        leave_ins.emp_code = emp_code
        leave_ins.emp_id = emp
        leave_ins.emp_name = user_name
        leave_ins.emp_email = user_email
        leave_ins.leave_type_id = leave_type_ins
        leave_ins.date_from = date_from
        leave_ins.date_to = date_to
        if sick_lev_doc_pic:
          leave_ins.sick_lev_doc_pic = uploaded_file_url
        leave_ins.leave_status = 'Approval Pending'
        leave_ins.save()
        # To Create Action Logs
        log_obj = LeaveAttendanceLogCl(leave_ins.id)
        action="Employee Leave Application Submitted"
        action_by = request.user.username
        log_obj.create_log(action,action_by)

        
        emp_obj = EmployeeMaster.objects.filter(emp_code=emp_code)
        for emp in emp_obj:
          if emp:
            reporting_manager_email = emp.reporting_authority_email
            mto = reporting_manager_email
            leave_app_mail(request,leave_ins,mto)
            msg = "Mail has been Sent Successfully to Reporting Manager !!!"
            return render(request, 'attendanceapp/thank_you.html',{'msg':msg})
          else:
            pass
    else:
      msg = "Selected Employee does not exists in Database"
      return render(request, 'attendanceapp/error.html',{'msg':msg})

  # except Exception as e: 
  #   return redirect('index')  



def edit_leave_application(request, leave_id):
  """
  To View Leave Application of Respective Employee 
  """
  try:
    leave = LeaveMaster.objects.get(id = leave_id)
    context = {
        'emp_code':leave.emp_code,
        'emp_name':leave.emp_name,
        'emp_email':leave.emp_email,
        'leave_type':leave.leave_type_id.leave_type,
        'date_from':leave.date_from,
        'date_to':leave.date_to,
        'status':leave.leave_status,
        'leave_id':leave.id
        }
    return render(request, 'attendanceapp/edit_leave_application.html', context)
  except Exception as e: 
    return redirect('leave_index')

def cancel_leave_application(request, emp_code, leave_id):
  """
  To Cancel Leave Application of Respective Employee 
  """
  try:
    leave = LeaveMaster.objects.get(id = leave_id)
    leave_type = leave.leave_type_id
    date_from = leave.date_from
    date_to = leave.date_to

    
    if leave.leave_status != 'Leave Cancelled':
      d1 = datetime.datetime.strptime(str(date_from).split(' ')[0], '%Y-%m-%d').date()            # From date of Attendance Report
      d2 = datetime.datetime.strptime(str(date_to).split(' ')[0], '%Y-%m-%d').date()            # Till date of Attendance Report
      delta = d2 - d1                                              # timedelta
      bal_lev_obj = BalancedLeaveMaster.objects.filter(emp_code=emp_code,balanced_leave_type_id=leave_type)
      for obj in bal_lev_obj :
        if obj:
          obj.balanced_leave = int(str(obj.balanced_leave)) + int(str(delta.days + 1))
          obj.save()
          
        else:
          msg = "Balanced Leave for Employee with given Leave Type does not exists"
          return render(request, 'attendanceapp/error.html',{'msg':msg})
    else:
      leave.leave_status == 'Leave Cancelled'
      msg = "These Leave Application has been Cancelled Already !!!"
      return render(request, 'attendanceapp/error.html',{'msg':msg})
    # To Create Action Logs
    log_obj = LeaveAttendanceLogCl(leave.id)
    action="Employee Leave Application has been Cancelled"
    action_by = request.user.username
    log_obj.create_log(action,action_by)

    bal_lev = BalancedLeaveMaster()
    context = {
      'emp_code':leave.emp_code,
      'emp_name':leave.emp_name,
      'emp_email':leave.emp_email,
      'leave_type':leave.leave_type_id.leave_type,
      'date_from':leave.date_from,
      'date_to':leave.date_to,
      'status':leave.leave_status,
      'leave_id':leave.id
      }
    return render(request, 'attendanceapp/edit_leave_application.html', context)
  except Exception as e: 
    return redirect('leave_index')



# Leave log list view 
# @login_required
def leave_logs_list(request,leave_id):

  # try:
  lev_ins=LeaveMaster.objects.get(id=leave_id)

  log_list = LeaveAttendanceLog.objects.filter(lev_id=lev_ins)

  for i in log_list:
    print(i.id)

  context={

  'log_list':log_list,
  'emp_code':lev_ins.emp_code,
  'emp_name':lev_ins.emp_name,
  'emp_email':lev_ins.emp_email,
  'leave_type':lev_ins.leave_type_id.leave_type,
  'date_from':lev_ins.date_from,
  'date_to':lev_ins.date_to,
  'status':lev_ins.leave_status,
  'leave_id':lev_ins.id
  
  }

  return render(request, 'attendanceapp/leave_logs_list.html', context)


  # except Exception as e: 
  #   print("Exception :",e)
  #   return redirect('index')


def add_leave_type(request):

  """
  To Add Holiday Type 
  """
  try:
    # states = StateName.objects.all()
    leave_type = LeaveTypesMaster.objects.all()

    return render(request, 'attendanceapp/add_leave_type.html',{'leave_type':leave_type})
  except Exception as e: 
    return redirect('leave_index')

def submit_leave_type(request):
  """
  To Submit Leave Type in Database
  """
  try:
    if request.method=='POST':
      leave_type = LeaveTypesMaster()
      lev_type = request.POST.get('lev_type',None)
      if lev_type is not None:
        leave_type.leave_type = lev_type
        leave_type.save()

      # msg = 'Leave Type added Successfully !!!'
      return render(request, 'attendanceapp/add_leave_type.html')
  except Exception as e: 
    return redirect('leave_index') 

def add_balanced_leave(request):
  """
  To ADD Balance Leave Of Users
  """
  try:
    if request.method=='POST':
      leave_type = LeaveTypesMaster.objects.all()
      balanced_leave_type = BalancedLeaveMaster.objects.all()
      lev_type = request.POST.get('lev_type',None)
      if lev_type is not None:
        leave_type.leave_type = lev_type
        leave_type.save()

      # msg = 'Leave Type added Successfully !!!'
      return render(request, 'attendanceapp/view_balanced_leave.html', {'leave_types': leave_type,'balanced_leave_types':balanced_leave_type})
  except Exception as e: 
    return redirect('leave_index')

def view_balanced_leave(request):
  """
  To View Balance Leave Of Users
  """
  bal_dic = {}             #Dictionary to add all balanced leave data of User
  bal_lev_dic = {}
  bal_list = []            #List to add all balanced data of User and pass it to html
  bal_lev_list = []
  # try:
  leave_type = LeaveTypesMaster.objects.all()
  balanced_leave_type_all = BalancedLeaveMaster.objects.all()
  balanced_leave_type = BalancedLeaveMaster.objects.values('emp_code').distinct()
  # print('-------balanced_leave_type',balanced_leave_type)
  for balance in balanced_leave_type:
    bal_obj = BalancedLeaveMaster.objects.filter(emp_code=balance.get('emp_code'))
    for bal in bal_obj:
      bal_dic = {'emp_code':bal.emp_code,'emp_name':bal.emp_name}
      for balance_all in balanced_leave_type_all:   #Getting all employees balanced leaves
        if bal.emp_code == balance_all.emp_code: #Codition to Get all leave_type of particular employee 
          for leave in leave_type:
            if balance_all.balanced_leave_type_id.leave_type == leave.leave_type:
              bal_lev_dic[bal.balanced_leave_type_id.short_name_leave_type] = bal.balanced_leave
    bal_lev_list.append(bal_lev_dic)
    bal_dic['bal_leave_no'] = bal_lev_list
    bal_lev_dic = {}
    bal_lev_list= []
        
    bal_list.append(bal_dic)
    bal_dic = {}
  return render(request, 'attendanceapp/view_balanced_leave.html', {'leave_types': leave_type,'bal_lists':bal_list})
  # except Exception as e: 
  #   return redirect('leave_index')

def leave_approval(request, emp_code, leave_id):
  """
  to send employee details to manager to approve/reject employee
  """
  try:
    dic = {}
    bal_lev_list = []
    emp = EmployeeMaster.objects.get(emp_code = emp_code)
    leave = LeaveMaster.objects.get(emp_code = emp_code, id = leave_id)
    leave_type = LeaveTypesMaster.objects.all()
    balanced_leave = BalancedLeaveMaster.objects.filter(emp_code = emp_code)
    for bal_lev in balanced_leave:
      dic[bal_lev.balanced_leave_type_id.short_name_leave_type] = bal_lev.balanced_leave
      bal_lev_list.append(dic)
      dic = {}
    # update_bal_lev = BalancedLeaveMaster.objects.get(balanced_leave_type_id = leave.leave_type_id)
    # if update_bal_lev:
    #   update_bal_lev.balanced_leave = update_bal_lev.balanced_leave - 2
    return render(request, 'attendanceapp/leave_approval.html', {'emp':emp,'bal_lev_lists':bal_lev_list,'leave_types':leave_type,'leave': leave,'date_from':leave.date_from})

  except Exception as e: 
    return redirect('index')

def approved_leave(request, emp_code, leave_id):
  """
  Function to Send mail to Employee after manager approves its Leave Request
  """
  try:
    if request.method == 'POST':
      emp = EmployeeMaster.objects.get(emp_code = emp_code)
      leave = LeaveMaster.objects.get(emp_code = emp_code, id = leave_id)
      lwp = request.POST.get('lwp',False)
      
      if lwp == 'on':
        leave_type = LeaveTypesMaster.objects.get(leave_type = 'LWP')
        leave.leave_type_id = leave_type
        leave.leave_status = 'Approved with LWP'
        leave.save()
      else:
        leave.leave_status = 'Approved'
        leave.save()
      d1 = datetime.datetime.strptime(str(leave.date_from).split(' ')[0], '%Y-%m-%d').date()            # From date of Attendance Report
      d2 = datetime.datetime.strptime(str(leave.date_to).split(' ')[0], '%Y-%m-%d').date()            # Till date of Attendance Report
      delta = d2 - d1                                              # timedelta

      bal_lev_obj = BalancedLeaveMaster.objects.filter(emp_code=emp_code,balanced_leave_type_id=leave.leave_type_id)
      for obj in bal_lev_obj :
        if obj:
          obj.balanced_leave = int(str(obj.balanced_leave)) - int(str(delta.days + 1))
          obj.save()
          
        else:
          msg = "Balanced Leave for Employee with given Leave Type does not exists"
          return render(request, 'attendanceapp/error.html',{'msg':msg})
      # if leave:
      #   leave.leave_status = 'Approved'
      #   leave.save()
      mto = leave.emp_email
      # mto = 'priyanka.chaurasia@rentokil-pci.com'
      print(mto)

      leave_mail_approved(request,leave,mto)
      msg = "Mail has been Sent Successfully !!!"
      return render(request, 'attendanceapp/thank_you.html',{'msg':msg})

  except Exception as e: 
    return redirect('index')


def reject_leave(request,leave_id,leave_emp_code):
  """
  to send reject employee email to service now and update emp status
  """
  try:
    if request.method == 'POST':

      emp_code = request.POST['emp_code']
      reason = request.POST['reason']
      leave_id = leave_id
      # emp = EmployeeMaster.objects.get(emp_code = leave_emp_code)
      leave = LeaveMaster.objects.get(emp_code = leave_emp_code, id = leave_id)
      if leave:
        leave.leave_status = 'Rejected'
        leave.save()
      msub="Leave Request Disapproved by Manager"
      # mbody="This mail is to inform that Employee : {}, is rejected by Branch Manager {}. Reason: {}.".format(leave.emp_name,leave.emp_id.emp_manager.bm_name,reason)
      mbody = """<html><body>
    <h3><p>Hello,</p></h3>
    <h3><p>Your Leave Application is Disapproved by Branch Manager Due to following reason :.</p></h3>
    """ + reason  + """
    <h3><p>Thanks and Regards,</p></h3>
    </body></html>"""

      mto = leave.emp_email
      # mto = "priyanka.chaurasia@rentokil-pci.com"

      mfrom='omiharjani@gmail.com'
      app_task(args=(msub,mbody,mfrom,mto))
    else :
      pass  
    msg = "Mail has been Sent Successfully !!!"
    return render(request, 'attendance/thank_you.html',{'msg',msg})

  except Exception as e: 
    print("Exception :",e)
    return redirect('index')  

def validate_send(request):
  """
  to check if emp_code already exist in database or not
  unique emp_code validation front end
  """
  try:
    emp_code = request.GET.get('emp_code', None)
    user_name = request.GET.get('user_name', None)
    date_from = request.GET.get('date_from', None)
    date_to = request.GET.get('date_to', None)
    leave_type = request.GET.get('leave_type', None)
    leave_type_ins = LeaveTypesMaster.objects.get(id=leave_type)
    data = {
      'is_taken': LeaveMaster.objects.filter(emp_code=emp_code,emp_name=user_name,date_from=date_from,date_to=date_to,leave_type_id=leave_type_ins).exists()
    }
    # data={'abc':hgh,''}
    return JsonResponse(data)
  except Exception as e:
    print("Exception :",e)
    return redirect('index')


def emp_code_validate(request):
  """
  to check if emp_code is present in database or not
  unique emp_code validation front end
  """
  try:
   
    emp_code = request.GET.get('emp_code', None)
    s = EmployeeMaster.objects.filter(emp_code=emp_code)
    if not s:
      data = {
        'is_taken': True
      }
      # data={'abc':hgh,''}
      return JsonResponse(data)
  except Exception as e:
    print("Exception :",e)
    return redirect('index')